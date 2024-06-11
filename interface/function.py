import os
from sheet import Sheet
from openpyxl import Workbook, load_workbook

SHEET_PATH = f'{os.path.expanduser("~")}/AppData/Roaming'
def verification():
    if not os.path.exists(f'{SHEET_PATH}/Sheets'):
        os.mkdir(f'{SHEET_PATH}/Sheets')

def verify_link(link: str):
    if link.count('kabum'): return True
    
def config_file(link: str):
    if not os.path.exists('config.txt'):
        with open('config.txt','a+') as file:
            file.write(f'{link}\n')
    else:
        with open('config.txt', 'r') as file:
            lines = file.readlines()
        
        links = [l[:-1] for l in lines]

        if not link in links: 
            with open('config.txt', 'a+') as file:
                file.write(f'{link}\n')
                
def config_sheet_file(link: str, name: str):
    filename = f'{SHEET_PATH}/Sheets/links.xlsx'

    caracters = ['\'', '/', ':', '*', '?', '"', '<', '>', '|']
    for caracter in caracters:
        if caracter in name:
            name = name.replace(caracter, "")  

    workbook = Workbook()
    links = [link, name]

    if not os.path.exists(filename):
        workbook.save(filename)
        
    wb = load_workbook(filename)
    ws = wb.active

    if not list(wb['Sheet'].values):
        wb['Sheet'].append(['Link', 'Name'])
        wb.save(filename)

    check_list = list()
    for value in wb["Sheet"].values:
        if not value[0] == "Link":
            check_list.append(value[0])

    if not link in check_list:
        ws.append(links)

    wb.save(filename)
    wb.close()

def remove_link(name:str) -> None:
    filename = f'{SHEET_PATH}/sheets/links.xlsx'

    wb = load_workbook(filename)
    for i, value in enumerate(wb["Sheet"].values):
        if value[1] == name:
            wb['Sheet'].delete_rows(i+1)
            break

    wb.save(filename)
    wb.close()
        
def get_links_sheets_file() -> list:
    filename = f'{SHEET_PATH}/sheets/links.xlsx'
    links = []

    wb = load_workbook(filename)
    wb.active
    
    for row in wb['Sheet'].values:
        if not row[0] == "Link":
            links.append(row[0])

    return links
                 
def monitor():
    with open("config.txt", 'r') as file:
        lines = file.readlines()
    
    return lines

def check_file_sheets():
    files = os.listdir(f'{SHEET_PATH}/sheets')
    files = [f for f in files if not f == "links.xlsx" in files and f.endswith(".xlsx")]
    return files