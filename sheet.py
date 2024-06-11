from openpyxl import Workbook, load_workbook
from datetime import date
import os

class Sheet():
    def __init__(self, fileName: str):
        self.filename = f'{self.format_filename(fileName)}.xlsx'
        self.path = f'{os.path.expanduser("~")}/AppData/Roaming'

        self.workbook = self.load_workbook()
        self.create_dir()
        self.check_file_execel()
     
    def format_filename(self, fileName: str):
        caracters = ['\'', '/', ':', '*', '?', '"', '<', '>', '|']
        for caracter in caracters:
            if caracter in fileName:
                fileName = fileName.replace(caracter, "")  
                
        return fileName
    
    def create_dir(self):
        if not os.path.exists(f'{self.path}/Sheets'):
            os.mkdir(f'{self.path}/Sheets')
    
    def check_file_execel(self):
        if not os.path.exists(f'{self.path}/Sheets/{self.filename}'):
            ws = self.workbook.active
            ws.append(['Nome', 'Data', 'Preço', 'A Vista'])
            self.save_workbook()
    
    def load_workbook(self):
        if not os.path.exists(f'{self.path}/Sheets/{self.filename}'): return Workbook()
        return load_workbook(f'{self.path}/Sheets/{self.filename}') 

    def save_workbook(self):
        self.workbook.save(f'{self.path}/Sheets/{self.filename}')
    
    def create_rows(self, data: list):
        ws = self.workbook.active
        ws.append(data)
        self.save_workbook()
        
    def check(self, data: list):        
        today = date.today()
        day = today.strftime("%d/%m/%Y")   
        last_row = len(range(self.workbook['Sheet'].max_row))
        
        # VERIFICA SE EXISTE VALORES IGUAIS PARA NÃO REPETIR.        
        if not self.workbook['Sheet'][f'B{last_row}'].value == day:
            self.create_rows(data)
            
        if self.workbook['Sheet'][f'B{last_row}'].value == day and not self.workbook['Sheet'][f'C{last_row}'].value == data[2] and not self.workbook['Sheet'][f'D{last_row}'].value == data[3]: 
            self.create_rows(data)